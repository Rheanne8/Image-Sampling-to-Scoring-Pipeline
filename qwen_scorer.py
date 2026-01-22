"""
Image Scoring Automation with Qwen via OpenRouter API
Scores sampled perspective images and exports to Excel with view grouping.
Best image per view is highlighted with bold formatting.
"""

from openai import OpenAI
from pathlib import Path
import pandas as pd
import json
import re
import time
import os
import base64
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Optional, Tuple, Dict, List
from io import BytesIO
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# ============== CONFIGURATION ==============
# Load API key from environment or set directly
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY", "YOUR_API_KEY_HERE")
IMAGE_FOLDER = r"C:\Users\rhean\Documents\RHEANNE\Acads\Y4S1\FYP\Final Pipeline\sampled_images"
OUTPUT_FILE = "perspective_scores_qwen.xlsx"
MODEL_NAME = "qwen/qwen-2.5-vl-7b-instruct:free"
# ===========================================

EVALUATION_PROMPT = """You are a conservative photographic evaluator.

You prioritise clarity, coherence, and purposeful visual decisions.

High scores are rare and require strong justification.

Briefly describe each image objectively.
Do not evaluate or praise.
Only note visible elements, main subject placement, horizon position, perspective, and spatial relationships.
A precision of 0.5 is allowed for scoring
Each image will receive at least half of the full marks allocated to each component.

Subject Emphasis: Assess how clearly and effectively the image directs attention to its intended subject
– Is the subject immediately identifiable?
– Does filling the frame strengthen focus or feel cramped?
– Do lines, framing, or depth guide attention toward the subject?
– Are there distractions that compete for attention?
Score from 0–2 based on clarity and strength of subject emphasis.

Viewpoint Creativity: Assess whether the camera viewpoint feels intentional and contributes meaningfully to how the subject is perceived.
– Does the camera angle, distance, or height offer a perspective that adds interest or insight?
– Does the placement of the subject within the frame feel deliberate rather than default?
– Does the framing choice reveal something that a neutral viewpoint would not?
Score from 0–2 based on how distinctive and purposeful the viewpoint feels.

Compositional Balance: Assess how visual weight is distributed across the frame and whether the composition feels resolved.
– Does the placement of major elements create stability or tension?
– Does the horizon placement support the composition or divide it awkwardly?
– Are multiple points of interest working together or competing?
– Is one side of the frame visually dominant without purpose?
Score from 0–2 based on how stable and coherent the overall balance feels.

Spatial Harmonization: Assess how well depth, perspective, and spatial relationships are integrated into a coherent visual space.
– Does the perspective create a clear sense of depth or spatial flow?
– Are foreground, midground, and background relationships clear and intentional?
– Do scale and distance between elements feel consistent?
– Does the space feel unified or fragmented?
Score from 0–4 based on how cohesive and readable the spatial structure is.

Sum the component scores for a total out of 10.
Briefly explain the rationale behind each component's score.
Briefly explain which component most influenced the final score.

IMPORTANT: At the end of your response, you MUST include a JSON block in exactly this format:
```json
{
    "subject_emphasis": {"score": <number>, "justification": "<brief text>"},
    "viewpoint_creativity": {"score": <number>, "justification": "<brief text>"},
    "compositional_balance": {"score": <number>, "justification": "<brief text>"},
    "spatial_harmonization": {"score": <number>, "justification": "<brief text>"},
    "total": <number>,
    "most_influential_component": "<component name>"
}
```
"""


# =============================================================================
# Filename Parsing
# =============================================================================

def parse_filename(filename: str) -> Optional[Tuple[str, int, int]]:
    """
    Parse a perspective image filename to extract components.

    Filename format: <place_name>_<view_id>_<perspective_id>.jpg

    Args:
        filename: The image filename (e.g., "museum_of_history_16k_5_3.jpg")

    Returns:
        Tuple of (place_name, view_id, perspective_id) or None if parsing fails
    """
    name = os.path.splitext(filename)[0]
    pattern = r'^(.+)_(\d+)_(\d+)$'
    match = re.match(pattern, name)

    if match:
        place_name = match.group(1)
        view_id = int(match.group(2))
        perspective_id = int(match.group(3))
        return (place_name, view_id, perspective_id)

    return None


# =============================================================================
# Image Discovery
# =============================================================================

def parse_subfolder_name(subfolder_name: str) -> Optional[Tuple[str, int]]:
    """
    Parse a subfolder name to extract location and view_id.

    Subfolder format: <place_name>_<view_id>
    Example: "museum_of_history_16k_7" -> ("museum_of_history_16k", 7)

    Returns:
        Tuple of (location, view_id) or None if parsing fails
    """
    # Match pattern: everything up to last underscore-separated number
    pattern = r'^(.+)_(\d+)$'
    match = re.match(pattern, subfolder_name)

    if match:
        location = match.group(1)
        view_id = int(match.group(2))
        return (location, view_id)

    return None


def get_sampled_images(folder: str) -> List[Dict]:
    """
    Get all image files from the sampled folder structure.

    Expected structure:
        sampled_images/
        ├── location1_viewid1/        # e.g., forest_16k_3
        │   ├── location1_viewid1_persp1.jpg
        │   ├── location1_viewid1_persp2.jpg
        │   └── ...
        └── location2_viewid2/
            └── ...

    Returns:
        List of dicts with keys: path, location, view_id, perspective_id, filename
    """
    folder_path = Path(folder)
    extensions = {'.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp'}
    images = []

    if not folder_path.exists():
        print(f"WARNING: Folder does not exist: {folder}")
        return images

    # Traverse subfolders (location_viewid combinations)
    for subfolder in folder_path.iterdir():
        if not subfolder.is_dir():
            continue

        subfolder_name = subfolder.name

        # Parse subfolder name to get location and view_id
        parsed_subfolder = parse_subfolder_name(subfolder_name)
        if parsed_subfolder:
            location, view_id = parsed_subfolder
        else:
            # Fallback: use subfolder name as location, view_id = -1
            location = subfolder_name
            view_id = -1

        # Get images in this subfolder
        for img_file in subfolder.iterdir():
            if not img_file.is_file():
                continue
            if img_file.suffix.lower() not in extensions:
                continue

            # Parse filename to get perspective_id
            parsed = parse_filename(img_file.name)
            if parsed:
                _, _, perspective_id = parsed
            else:
                perspective_id = -1

            images.append({
                'path': img_file,
                'location': location,
                'view_id': view_id,
                'perspective_id': perspective_id,
                'filename': img_file.name
            })

    # Sort by location, then view_id, then perspective_id
    images.sort(key=lambda x: (x['location'], x['view_id'], x['perspective_id']))

    return images


# =============================================================================
# OpenRouter API Functions
# =============================================================================

def setup_openrouter(api_key: str):
    """Configure OpenRouter API client."""
    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
    )
    return client


def image_to_base64(image_path: Path) -> str:
    """Convert image to base64 string."""
    img = Image.open(image_path)

    # Convert to RGB if necessary (remove alpha channel)
    if img.mode in ('RGBA', 'LA', 'P'):
        background = Image.new('RGB', img.size, (255, 255, 255))
        if img.mode == 'P':
            img = img.convert('RGBA')
        background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
        img = background

    # Resize if image is too large (optional, helps with API limits)
    max_size = 2048
    if max(img.size) > max_size:
        img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)

    # Convert to base64
    buffered = BytesIO()
    img.save(buffered, format="JPEG", quality=85)
    img_str = base64.b64encode(buffered.getvalue()).decode()

    return img_str


def extract_json_from_response(response_text: str) -> dict:
    """Extract JSON data from the model's response."""
    patterns = [
        r'```json\s*(.*?)\s*```',
        r'```\s*({\s*"subject_emphasis".*?})\s*```',
        r'```\s*(.*?)\s*```',
    ]

    for pattern in patterns:
        json_match = re.search(pattern, response_text, re.DOTALL)
        if json_match:
            try:
                parsed = json.loads(json_match.group(1))
                if "subject_emphasis" in parsed:
                    return parsed
            except json.JSONDecodeError:
                continue

    start_match = re.search(r'\{\s*"subject_emphasis"', response_text)
    if start_match:
        start_idx = start_match.start()
        brace_count = 0
        end_idx = start_idx
        for i, char in enumerate(response_text[start_idx:]):
            if char == '{':
                brace_count += 1
            elif char == '}':
                brace_count -= 1
                if brace_count == 0:
                    end_idx = start_idx + i + 1
                    break

        try:
            json_str = response_text[start_idx:end_idx]
            return json.loads(json_str)
        except json.JSONDecodeError:
            pass

    for match in re.finditer(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response_text):
        try:
            parsed = json.loads(match.group())
            if "subject_emphasis" in parsed or "total" in parsed:
                return parsed
        except json.JSONDecodeError:
            continue

    return {
        "subject_emphasis": {"score": None, "justification": "Parse error"},
        "viewpoint_creativity": {"score": None, "justification": "Parse error"},
        "compositional_balance": {"score": None, "justification": "Parse error"},
        "spatial_harmonization": {"score": None, "justification": "Parse error"},
        "total": None,
        "most_influential_component": "Parse error"
    }


def test_api_connection(client, model_name: str) -> bool:
    """Test the API connection with a simple request."""
    print("Testing API connection...")
    try:
        response = client.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "user", "content": "Say 'API working' in exactly 2 words."}
            ],
            max_tokens=50
        )
        response_text = response.choices[0].message.content.strip()
        print(f"  API test successful: {response_text[:50]}")
        return True
    except Exception as e:
        print(f"  API test FAILED!")
        print(f"  Error type: {type(e).__name__}")
        print(f"  Error message: {str(e)}")
        return False


def score_image(client, model_name: str, image_info: Dict, max_retries: int = 3) -> dict:
    """Score a single image using Qwen via OpenRouter with retry logic."""
    image_path = image_info['path']

    for attempt in range(max_retries):
        try:
            # Convert image to base64
            base64_image = image_to_base64(image_path)

            # Create the API request
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": EVALUATION_PROMPT
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                max_tokens=2000
            )

            response_text = response.choices[0].message.content

            data = extract_json_from_response(response_text)

            # Add metadata from image_info
            data['filename'] = image_info['filename']
            data['location'] = image_info['location']
            data['view_id'] = image_info['view_id']
            data['perspective_id'] = image_info['perspective_id']
            data['full_response'] = response_text

            return data

        except Exception as e:
            error_str = str(e)
            error_type = type(e).__name__

            # Print detailed error info for debugging
            print(f"         ERROR on attempt {attempt + 1}/{max_retries}")
            print(f"         Error type: {error_type}")
            print(f"         Error message: {error_str[:200]}")

            # Check for rate limiting (various patterns)
            is_rate_limit = any(x in error_str.lower() for x in ["429", "rate", "quota", "limit", "resource exhausted"])

            if is_rate_limit and attempt < max_retries - 1:
                wait_time = 60 * (attempt + 1)
                print(f"         Rate limited. Waiting {wait_time}s before retry...")
                time.sleep(wait_time)
                continue

            return {
                "filename": image_info['filename'],
                "location": image_info['location'],
                "view_id": image_info['view_id'],
                "perspective_id": image_info['perspective_id'],
                "subject_emphasis": {"score": None, "justification": f"Error: {error_str[:100]}"},
                "viewpoint_creativity": {"score": None, "justification": f"Error: {error_str[:100]}"},
                "compositional_balance": {"score": None, "justification": f"Error: {error_str[:100]}"},
                "spatial_harmonization": {"score": None, "justification": f"Error: {error_str[:100]}"},
                "total": None,
                "most_influential_component": "Error",
                "full_response": f"Error ({error_type}): {error_str}"
            }

    return {
        "filename": image_info['filename'],
        "location": image_info['location'],
        "view_id": image_info['view_id'],
        "perspective_id": image_info['perspective_id'],
        "subject_emphasis": {"score": None, "justification": "Max retries exceeded"},
        "viewpoint_creativity": {"score": None, "justification": "Max retries exceeded"},
        "compositional_balance": {"score": None, "justification": "Max retries exceeded"},
        "spatial_harmonization": {"score": None, "justification": "Max retries exceeded"},
        "total": None,
        "most_influential_component": "Max retries exceeded",
        "full_response": "Max retries exceeded"
    }


# =============================================================================
# Results Processing and Excel Export
# =============================================================================

def group_and_sort_results(results: List[dict]) -> List[dict]:
    """
    Group results by location+view_id and sort by total score within each group.
    Returns a flat list with groups sorted by score (best first within each view).
    """
    from collections import defaultdict

    # Group by (location, view_id)
    groups = defaultdict(list)
    for r in results:
        key = (r.get('location', ''), r.get('view_id', -1))
        groups[key].append(r)

    # Sort each group by total score (descending), None scores go last
    for key in groups:
        groups[key].sort(
            key=lambda x: (x.get('total') is None, -(x.get('total') or 0))
        )

    # Mark best in each view
    for key in groups:
        if groups[key]:
            groups[key][0]['is_best'] = True
            for item in groups[key][1:]:
                item['is_best'] = False

    # Flatten back to list, sorted by location then view_id
    sorted_keys = sorted(groups.keys())
    flat_results = []
    for key in sorted_keys:
        flat_results.extend(groups[key])
        # Add a None entry to represent blank row separator
        flat_results.append(None)

    # Remove trailing None
    if flat_results and flat_results[-1] is None:
        flat_results.pop()

    return flat_results


def results_to_dataframe(results: List[dict]) -> pd.DataFrame:
    """Convert results to a formatted DataFrame with view grouping."""
    # Group and sort results
    sorted_results = group_and_sort_results(results)

    rows = []
    for r in sorted_results:
        if r is None:
            # Blank row separator
            rows.append({
                "Location": "",
                "View ID": "",
                "Image": "",
                "Subject Emphasis (0-2)": "",
                "SE Justification": "",
                "Viewpoint Creativity (0-2)": "",
                "VC Justification": "",
                "Compositional Balance (0-2)": "",
                "CB Justification": "",
                "Spatial Harmonization (0-4)": "",
                "SH Justification": "",
                "Total (/10)": "",
                "Best in View": "",
                "_is_best": False,
                "_is_blank": True
            })
        else:
            rows.append({
                "Location": r.get("location", ""),
                "View ID": r.get("view_id", ""),
                "Image": r.get("filename", ""),
                "Subject Emphasis (0-2)": r.get("subject_emphasis", {}).get("score"),
                "SE Justification": r.get("subject_emphasis", {}).get("justification", ""),
                "Viewpoint Creativity (0-2)": r.get("viewpoint_creativity", {}).get("score"),
                "VC Justification": r.get("viewpoint_creativity", {}).get("justification", ""),
                "Compositional Balance (0-2)": r.get("compositional_balance", {}).get("score"),
                "CB Justification": r.get("compositional_balance", {}).get("justification", ""),
                "Spatial Harmonization (0-4)": r.get("spatial_harmonization", {}).get("score"),
                "SH Justification": r.get("spatial_harmonization", {}).get("justification", ""),
                "Total (/10)": r.get("total"),
                "Best in View": "✓" if r.get("is_best") else "",
                "_is_best": r.get("is_best", False),
                "_is_blank": False
            })

    return pd.DataFrame(rows)


def save_with_formatting(df: pd.DataFrame, output_file: str):
    """Save DataFrame to Excel with bold formatting for best images."""
    # Get the metadata columns
    is_best_col = df['_is_best'].tolist()
    is_blank_col = df['_is_blank'].tolist()

    # Remove metadata columns before saving
    df_export = df.drop(columns=['_is_best', '_is_blank'])

    # Save to Excel first
    df_export.to_excel(output_file, index=False, sheet_name="Scores")

    # Now apply formatting using openpyxl
    wb = load_workbook(output_file)
    ws = wb.active

    # Define styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    best_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green

    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Format data rows
    for row_idx, (is_best, is_blank) in enumerate(zip(is_best_col, is_blank_col), start=2):
        if is_blank:
            continue

        if is_best:
            for cell in ws[row_idx]:
                cell.font = bold_font
                cell.fill = best_fill

    # Adjust column widths
    column_widths = {
        'A': 20,   # Location
        'B': 8,    # View ID
        'C': 30,   # Image
        'D': 10,   # SE Score
        'E': 40,   # SE Justification
        'F': 10,   # VC Score
        'G': 40,   # VC Justification
        'H': 10,   # CB Score
        'I': 40,   # CB Justification
        'J': 10,   # SH Score
        'K': 40,   # SH Justification
        'L': 10,   # Total
        'M': 12,   # Best in View
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Enable text wrapping for justification columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_file)


# =============================================================================
# Main Entry Point
# =============================================================================

def main():
    print("=" * 60)
    print("Qwen Perspective Image Scorer (via OpenRouter)")
    print("=" * 60)

    # Setup
    if OPENROUTER_API_KEY == "YOUR_API_KEY_HERE":
        print("ERROR: Please set your API key!")
        print("Either:")
        print("  1. Set OPENROUTER_API_KEY in your .env file")
        print("  2. Replace YOUR_API_KEY_HERE in this script")
        print("\nGet your key at: https://openrouter.ai/keys")
        return

    print(f"API Key: {OPENROUTER_API_KEY[:10]}...{OPENROUTER_API_KEY[-4:]}")
    print(f"Model: {MODEL_NAME}")

    client = setup_openrouter(OPENROUTER_API_KEY)

    # Test API connection before proceeding
    if not test_api_connection(client, MODEL_NAME):
        print("\nAPI connection failed. Please check:")
        print("  1. Is your API key valid?")
        print("  2. Visit https://openrouter.ai/keys to verify your key")
        print("  3. Check https://openrouter.ai/docs for API documentation")
        return

    print("-" * 60)

    # Get images from sampled folder
    print(f"Scanning: {IMAGE_FOLDER}")
    images = get_sampled_images(IMAGE_FOLDER)

    if not images:
        print(f"ERROR: No images found in '{IMAGE_FOLDER}'")
        print("Please run the sampling pipeline first.")
        return

    # Count views
    views = set((img['location'], img['view_id']) for img in images)
    print(f"Found {len(images)} images across {len(views)} views")
    print("-" * 60)

    # Process each image
    results = []
    for i, image_info in enumerate(images, 1):
        print(f"[{i}/{len(images)}] Processing: {image_info['location']}/{image_info['filename']}")
        result = score_image(client, MODEL_NAME, image_info)
        results.append(result)

        # Show progress
        if result.get("total") is not None:
            print(f"         Score: {result['total']}/10")
        else:
            print(f"         Warning: Could not parse score")
            response_snippet = result.get('full_response', '')[:200].replace('\n', ' ')
            print(f"         Response preview: {response_snippet}...")

        # Rate limiting - adjust this based on OpenRouter's free tier limits
        if i < len(images):
            time.sleep(3)  # Conservative delay for free tier

        # Save partial results every 5 images
        if i % 5 == 0:
            df = results_to_dataframe(results)
            save_with_formatting(df, OUTPUT_FILE)
            print(f"         [Partial save: {i} images]")

    # Save final results with formatting
    df = results_to_dataframe(results)
    save_with_formatting(df, OUTPUT_FILE)

    # Print summary
    print("-" * 60)
    print("SUMMARY BY VIEW:")
    print("-" * 60)

    sorted_results = group_and_sort_results(results)
    current_view = None
    for r in sorted_results:
        if r is None:
            continue
        view_key = (r.get('location'), r.get('view_id'))
        if view_key != current_view:
            current_view = view_key
            print(f"\n{r.get('location')} - View {r.get('view_id')}:")

        marker = "  ★ BEST" if r.get('is_best') else "       "
        total = r.get('total', 'N/A')
        print(f"  {marker} {r.get('filename')}: {total}/10")

    print("\n" + "-" * 60)
    print(f"Results saved to: {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
