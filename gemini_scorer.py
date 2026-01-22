"""
Image Scoring Automation with Gemini API
Scores sampled perspective images and exports to Excel with view grouping.
Best image per view is highlighted with bold formatting.
"""

import google.generativeai as genai
from pathlib import Path
import pandas as pd
import json
import re
import time
import os
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from typing import Optional, Tuple, Dict, List

# ============== CONFIGURATION ==============
API_KEY = "AIzaSyBmJnZTOFo-Eq46q7Lv32YGkZQJtt19kBg"  # Replace with your Gemini API key
IMAGE_FOLDER = r"C:\Users\rhean\Documents\RHEANNE\Acads\Y4S1\FYP\Final Pipeline\sampled_images"
OUTPUT_FILE = "perspective_scores_gemini.xlsx"
MODEL_NAME = "gemini-2.5-flash"  # Free tier model
# ===========================================

EVALUATION_PROMPT = """You are a photographic evaluator who prioritises clarity, coherence, and purposeful visual decisions.

Briefly describe each image objectively.
Do not evaluate or praise.
Only note visible elements, main subject placement, horizon position, perspective, and spatial relationships.
A precision of 0.5 is allowed for scoring.

Subject Emphasis: Assess how clearly and effectively the image directs attention to its intended subject
– Is the subject immediately identifiable?
– Does filling the frame strengthen focus or feel cramped?
– Do lines, framing, or depth guide attention toward the subject?
– Are there distractions that compete for attention?
Score from 1–2 based on clarity and strength of subject emphasis.

Viewpoint Creativity: Assess whether the camera viewpoint feels intentional and contributes meaningfully to how the subject is perceived.
– Does the camera angle, distance, or height offer a perspective that adds interest or insight?
– Does the placement of the subject within the frame feel deliberate rather than default?
– Does the framing choice reveal something that a neutral viewpoint would not?
Score from 1–2 based on how distinctive and purposeful the viewpoint feels.

Compositional Balance: Assess how visual weight is distributed across the frame and whether the composition feels resolved.
– Does the placement of major elements create stability or tension?
– Does the horizon placement support the composition or divide it awkwardly?
– Are multiple points of interest working together or competing?
– Is one side of the frame visually dominant without purpose?
Score from 1–2 based on how stable and coherent the overall balance feels.

Spatial Harmonization: Assess how well depth, perspective, and spatial relationships are integrated into a coherent visual space.
– Does the perspective create a clear sense of depth or spatial flow?
– Are foreground, midground, and background relationships clear and intentional?
– Do scale and distance between elements feel consistent?
– Does the space feel unified or fragmented?
Score from 2–4 based on how cohesive and readable the spatial structure is.

Sum the component scores for a total out of 10.
Briefly explain the rationale behind each component's score.
Briefly explain which component most influenced the final score.

IMPORTANT: At the end of your response, you MUST include a JSON block in exactly this format:
```json
{
    "subject_emphasis": {"score": <number>, "justification": "<brief text>"},
    "viewpoint_creativity": {"score": <number>, "justification": "<brief text>"},
    "compositional_balance": {"score": <number>, "justification": "<brief text>"},
    "spatial_harmonization": {"score": <number>, "justification": "<brief text>"},
    "total": <number>,
    "most_influential_component": "<component name>"
}
```

CRITICAL SCORING RULE - YOU MUST FOLLOW THIS:
The MINIMUM score for each component is:
- Subject Emphasis: minimum 1.0 (out of 2)
- Viewpoint Creativity: minimum 1.0 (out of 2)
- Compositional Balance: minimum 1.0 (out of 2)
- Spatial Harmonization: minimum 2.0 (out of 4)
- Total: minimum 5.0 (out of 10)

NEVER score below these minimums. Every image receives at least half marks per component.
"""


# =============================================================================
# Filename Parsing
# =============================================================================

def parse_filename(filename: str) -> Optional[Tuple[str, int, int]]:
    """
    Parse a perspective image filename to extract components.

    Filename format: <place_name>_<view_id>_<perspective_id>.jpg

    Args:
        filename: The image filename (e.g., "museum_of_history_16k_5_3.jpg")

    Returns:
        Tuple of (place_name, view_id, perspective_id) or None if parsing fails
    """
    name = os.path.splitext(filename)[0]
    pattern = r'^(.+)_(\d+)_(\d+)$'
    match = re.match(pattern, name)

    if match:
        place_name = match.group(1)
        view_id = int(match.group(2))
        perspective_id = int(match.group(3))
        return (place_name, view_id, perspective_id)

    return None


# =============================================================================
# Image Discovery
# =============================================================================

def parse_subfolder_name(subfolder_name: str) -> Optional[Tuple[str, int]]:
    """
    Parse a subfolder name to extract location and view_id.

    Subfolder format: <place_name>_<view_id>
    Example: "museum_of_history_16k_7" -> ("museum_of_history_16k", 7)

    Returns:
        Tuple of (location, view_id) or None if parsing fails
    """
    # Match pattern: everything up to last underscore-separated number
    pattern = r'^(.+)_(\d+)$'
    match = re.match(pattern, subfolder_name)

    if match:
        location = match.group(1)
        view_id = int(match.group(2))
        return (location, view_id)

    return None


def get_sampled_images(folder: str) -> List[Dict]:
    """
    Get all image files from the sampled folder structure.

    Expected structure:
        sampled_images/
        ├── location1_viewid1/        # e.g., forest_16k_3
        │   ├── location1_viewid1_persp1.jpg
        │   ├── location1_viewid1_persp2.jpg
        │   └── ...
        └── location2_viewid2/
            └── ...

    Returns:
        List of dicts with keys: path, location, view_id, perspective_id, filename
    """
    folder_path = Path(folder)
    extensions = {'.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp'}
    images = []

    if not folder_path.exists():
        print(f"WARNING: Folder does not exist: {folder}")
        return images

    # Traverse subfolders (location_viewid combinations)
    for subfolder in folder_path.iterdir():
        if not subfolder.is_dir():
            continue

        subfolder_name = subfolder.name

        # Parse subfolder name to get location and view_id
        parsed_subfolder = parse_subfolder_name(subfolder_name)
        if parsed_subfolder:
            location, view_id = parsed_subfolder
        else:
            # Fallback: use subfolder name as location, view_id = -1
            location = subfolder_name
            view_id = -1

        # Get images in this subfolder
        for img_file in subfolder.iterdir():
            if not img_file.is_file():
                continue
            if img_file.suffix.lower() not in extensions:
                continue

            # Parse filename to get perspective_id
            parsed = parse_filename(img_file.name)
            if parsed:
                _, _, perspective_id = parsed
            else:
                perspective_id = -1

            images.append({
                'path': img_file,
                'location': location,
                'view_id': view_id,
                'perspective_id': perspective_id,
                'filename': img_file.name
            })

    # Sort by location, then view_id, then perspective_id
    images.sort(key=lambda x: (x['location'], x['view_id'], x['perspective_id']))

    return images


# =============================================================================
# Gemini API Functions
# =============================================================================

def setup_gemini(api_key: str):
    """Configure Gemini API."""
    genai.configure(api_key=api_key)
    return genai.GenerativeModel(MODEL_NAME)


def extract_json_from_response(response_text: str) -> dict:
    """Extract JSON data from the model's response."""
    patterns = [
        r'```json\s*(.*?)\s*```',
        r'```\s*({\s*"subject_emphasis".*?})\s*```',
        r'```\s*(.*?)\s*```',
    ]

    for pattern in patterns:
        json_match = re.search(pattern, response_text, re.DOTALL)
        if json_match:
            try:
                parsed = json.loads(json_match.group(1))
                if "subject_emphasis" in parsed:
                    return parsed
            except json.JSONDecodeError:
                continue

    start_match = re.search(r'\{\s*"subject_emphasis"', response_text)
    if start_match:
        start_idx = start_match.start()
        brace_count = 0
        end_idx = start_idx
        for i, char in enumerate(response_text[start_idx:]):
            if char == '{':
                brace_count += 1
            elif char == '}':
                brace_count -= 1
                if brace_count == 0:
                    end_idx = start_idx + i + 1
                    break

        try:
            json_str = response_text[start_idx:end_idx]
            return json.loads(json_str)
        except json.JSONDecodeError:
            pass

    for match in re.finditer(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response_text):
        try:
            parsed = json.loads(match.group())
            if "subject_emphasis" in parsed or "total" in parsed:
                return parsed
        except json.JSONDecodeError:
            continue

    return {
        "subject_emphasis": {"score": None, "justification": "Parse error"},
        "viewpoint_creativity": {"score": None, "justification": "Parse error"},
        "compositional_balance": {"score": None, "justification": "Parse error"},
        "spatial_harmonization": {"score": None, "justification": "Parse error"},
        "total": None,
        "most_influential_component": "Parse error"
    }


def test_api_connection(model) -> bool:
    """Test the API connection with a simple request."""
    print("Testing API connection...")
    try:
        response = model.generate_content("Say 'API working' in exactly 2 words.")
        print(f"  API test successful: {response.text.strip()[:50]}")
        return True
    except Exception as e:
        print(f"  API test FAILED!")
        print(f"  Error type: {type(e).__name__}")
        print(f"  Error message: {str(e)}")
        return False


def score_image(model, image_info: Dict, max_retries: int = 3) -> dict:
    """Score a single image using Gemini with retry logic."""
    image_path = image_info['path']

    for attempt in range(max_retries):
        try:
            img = Image.open(image_path)
            response = model.generate_content([EVALUATION_PROMPT, img])
            response_text = response.text

            data = extract_json_from_response(response_text)

            # Add metadata from image_info
            data['filename'] = image_info['filename']
            data['location'] = image_info['location']
            data['view_id'] = image_info['view_id']
            data['perspective_id'] = image_info['perspective_id']
            data['path'] = image_info['path']
            data['full_response'] = response_text

            return data

        except Exception as e:
            error_str = str(e)
            error_type = type(e).__name__

            # Print detailed error info for debugging
            print(f"         ERROR on attempt {attempt + 1}/{max_retries}")
            print(f"         Error type: {error_type}")
            print(f"         Error message: {error_str[:200]}")

            # Check for rate limiting (various patterns)
            is_rate_limit = any(x in error_str.lower() for x in ["429", "rate", "quota", "limit", "resource exhausted"])

            if is_rate_limit and attempt < max_retries - 1:
                wait_time = 60 * (attempt + 1)
                print(f"         Rate limited. Waiting {wait_time}s before retry...")
                time.sleep(wait_time)
                continue

            return {
                "filename": image_info['filename'],
                "location": image_info['location'],
                "view_id": image_info['view_id'],
                "perspective_id": image_info['perspective_id'],
                "path": image_info['path'],
                "subject_emphasis": {"score": None, "justification": f"Error: {error_str[:100]}"},
                "viewpoint_creativity": {"score": None, "justification": f"Error: {error_str[:100]}"},
                "compositional_balance": {"score": None, "justification": f"Error: {error_str[:100]}"},
                "spatial_harmonization": {"score": None, "justification": f"Error: {error_str[:100]}"},
                "total": None,
                "most_influential_component": "Error",
                "full_response": f"Error ({error_type}): {error_str}"
            }

    return {
        "filename": image_info['filename'],
        "location": image_info['location'],
        "view_id": image_info['view_id'],
        "perspective_id": image_info['perspective_id'],
        "path": image_info['path'],
        "subject_emphasis": {"score": None, "justification": "Max retries exceeded"},
        "viewpoint_creativity": {"score": None, "justification": "Max retries exceeded"},
        "compositional_balance": {"score": None, "justification": "Max retries exceeded"},
        "spatial_harmonization": {"score": None, "justification": "Max retries exceeded"},
        "total": None,
        "most_influential_component": "Max retries exceeded",
        "full_response": "Max retries exceeded"
    }


# =============================================================================
# Results Processing and Excel Export
# =============================================================================

def group_and_sort_results(results: List[dict]) -> List[dict]:
    """
    Group results by location+view_id and sort by total score within each group.
    Returns a flat list with groups sorted by score (best first within each view).
    """
    from collections import defaultdict

    # Group by (location, view_id)
    groups = defaultdict(list)
    for r in results:
        key = (r.get('location', ''), r.get('view_id', -1))
        groups[key].append(r)

    # Sort each group by total score (descending), None scores go last
    for key in groups:
        groups[key].sort(
            key=lambda x: (x.get('total') is None, -(x.get('total') or 0))
        )

    # Mark best in each view
    for key in groups:
        if groups[key]:
            groups[key][0]['is_best'] = True
            for item in groups[key][1:]:
                item['is_best'] = False

    # Flatten back to list, sorted by location then view_id
    sorted_keys = sorted(groups.keys())
    flat_results = []
    for key in sorted_keys:
        flat_results.extend(groups[key])
        # Add a None entry to represent blank row separator
        flat_results.append(None)

    # Remove trailing None
    if flat_results and flat_results[-1] is None:
        flat_results.pop()

    return flat_results


def results_to_dataframe(results: List[dict]) -> pd.DataFrame:
    """Convert results to a formatted DataFrame with view grouping."""
    # Group and sort results
    sorted_results = group_and_sort_results(results)

    rows = []
    for r in sorted_results:
        if r is None:
            # Blank row separator
            rows.append({
                "Location": "",
                "View ID": "",
                "Image ID": "",
                "Image Preview": "",  # Placeholder for actual image
                "Subject Emphasis (0-2)": "",
                "SE Justification": "",
                "Viewpoint Creativity (0-2)": "",
                "VC Justification": "",
                "Compositional Balance (0-2)": "",
                "CB Justification": "",
                "Spatial Harmonization (0-4)": "",
                "SH Justification": "",
                "Total (/10)": "",
                "Best in View": "",
                "_is_best": False,
                "_is_blank": True,
                "_image_path": None
            })
        else:
            rows.append({
                "Location": r.get("location", ""),
                "View ID": r.get("view_id", ""),
                "Image ID": r.get("filename", ""),
                "Image Preview": "",  # Will be filled with actual image
                "Subject Emphasis (0-2)": r.get("subject_emphasis", {}).get("score"),
                "SE Justification": r.get("subject_emphasis", {}).get("justification", ""),
                "Viewpoint Creativity (0-2)": r.get("viewpoint_creativity", {}).get("score"),
                "VC Justification": r.get("viewpoint_creativity", {}).get("justification", ""),
                "Compositional Balance (0-2)": r.get("compositional_balance", {}).get("score"),
                "CB Justification": r.get("compositional_balance", {}).get("justification", ""),
                "Spatial Harmonization (0-4)": r.get("spatial_harmonization", {}).get("score"),
                "SH Justification": r.get("spatial_harmonization", {}).get("justification", ""),
                "Total (/10)": r.get("total"),
                "Best in View": "✓" if r.get("is_best") else "",
                "_is_best": r.get("is_best", False),
                "_is_blank": False,
                "_image_path": str(r.get("path", "")) if r.get("path") else None
            })

    return pd.DataFrame(rows)


def save_with_formatting(df: pd.DataFrame, output_file: str):
    """Save DataFrame to Excel with bold formatting for best images and embedded image previews."""
    # Get the metadata columns
    is_best_col = df['_is_best'].tolist()
    is_blank_col = df['_is_blank'].tolist()
    image_paths = df['_image_path'].tolist()

    # Remove metadata columns before saving
    df_export = df.drop(columns=['_is_best', '_is_blank', '_image_path'])

    # Save to Excel first
    df_export.to_excel(output_file, index=False, sheet_name="Scores")

    # Now apply formatting using openpyxl
    wb = load_workbook(output_file)
    ws = wb.active

    # Define styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    best_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green

    # Image thumbnail settings
    IMG_WIDTH = 150  # pixels
    IMG_HEIGHT = 100  # pixels
    ROW_HEIGHT_WITH_IMAGE = 80  # Excel row height units (approx 100 pixels)

    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Format data rows and insert images
    for row_idx, (is_best, is_blank, img_path) in enumerate(zip(is_best_col, is_blank_col, image_paths), start=2):
        if is_blank:
            continue

        if is_best:
            for cell in ws[row_idx]:
                cell.font = bold_font
                cell.fill = best_fill

        # Insert image preview in column D (Image Preview)
        if img_path:
            img_path_obj = Path(img_path)
            if img_path_obj.exists():
                try:
                    # Create thumbnail
                    pil_img = Image.open(img_path_obj)
                    pil_img.thumbnail((IMG_WIDTH, IMG_HEIGHT), Image.Resampling.LANCZOS)

                    # Save thumbnail to temp file using absolute path
                    output_dir = Path(output_file).resolve().parent
                    temp_dir = output_dir / ".temp_thumbnails"
                    temp_dir.mkdir(exist_ok=True)
                    temp_path = temp_dir / f"thumb_{row_idx}.png"
                    pil_img.save(str(temp_path), "PNG")
                    pil_img.close()

                    # Insert image into Excel
                    xl_img = XLImage(str(temp_path))
                    xl_img.width = IMG_WIDTH
                    xl_img.height = IMG_HEIGHT

                    # Anchor image to cell D{row_idx}
                    cell_ref = f"D{row_idx}"
                    ws.add_image(xl_img, cell_ref)

                    # Set row height to accommodate image
                    ws.row_dimensions[row_idx].height = ROW_HEIGHT_WITH_IMAGE

                except Exception as e:
                    print(f"Warning: Could not embed image for row {row_idx}: {type(e).__name__}: {e}")
            else:
                print(f"Warning: Image path does not exist for row {row_idx}: {img_path}")

    # Adjust column widths
    column_widths = {
        'A': 20,   # Location
        'B': 8,    # View ID
        'C': 30,   # Image ID
        'D': 22,   # Image Preview (width for ~150px image)
        'E': 10,   # SE Score
        'F': 40,   # SE Justification
        'G': 10,   # VC Score
        'H': 40,   # VC Justification
        'I': 10,   # CB Score
        'J': 40,   # CB Justification
        'K': 10,   # SH Score
        'L': 40,   # SH Justification
        'M': 10,   # Total
        'N': 12,   # Best in View
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Enable text wrapping for justification columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_file)
    print(f"Excel file saved with {len([p for p in image_paths if p])} image previews")

    # Clean up temporary thumbnail files
    output_dir = Path(output_file).resolve().parent
    temp_dir = output_dir / ".temp_thumbnails"
    if temp_dir.exists():
        try:
            import shutil
            shutil.rmtree(temp_dir)
        except Exception:
            pass  # Ignore cleanup errors


# =============================================================================
# Main Entry Point
# =============================================================================

def main():
    print("=" * 60)
    print("Gemini Perspective Image Scorer")
    print("=" * 60)

    # Setup
    if API_KEY == "YOUR_API_KEY_HERE":
        print("ERROR: Please set your API key in the script!")
        print("Open gemini_scorer.py and replace YOUR_API_KEY_HERE with your key.")
        return

    print(f"API Key: {API_KEY[:10]}...{API_KEY[-4:]}")
    print(f"Model: {MODEL_NAME}")

    model = setup_gemini(API_KEY)

    # Test API connection before proceeding
    if not test_api_connection(model):
        print("\nAPI connection failed. Please check:")
        print("  1. Is your API key valid?")
        print("  2. Is the model name correct? Try 'gemini-1.5-flash' or 'gemini-pro'")
        print("  3. Do you have billing enabled on your Google Cloud project?")
        print("  4. Check https://aistudio.google.com/ for your API usage")
        return

    print("-" * 60)

    # Get images from sampled folder
    print(f"Scanning: {IMAGE_FOLDER}")
    images = get_sampled_images(IMAGE_FOLDER)

    if not images:
        print(f"ERROR: No images found in '{IMAGE_FOLDER}'")
        print("Please run the sampling pipeline first.")
        return

    # Count views
    views = set((img['location'], img['view_id']) for img in images)
    print(f"Found {len(images)} images across {len(views)} views")
    print("-" * 60)

    # Process each image
    results = []
    for i, image_info in enumerate(images, 1):
        print(f"[{i}/{len(images)}] Processing: {image_info['location']}/{image_info['filename']}")
        result = score_image(model, image_info)
        results.append(result)

        # Show progress
        if result.get("total") is not None:
            print(f"         Score: {result['total']}/10")
        else:
            print(f"         Warning: Could not parse score")
            response_snippet = result.get('full_response', '')[:200].replace('\n', ' ')
            print(f"         Response preview: {response_snippet}...")

        # Rate limiting
        if i < len(images):
            time.sleep(6)

        # Save partial results every 5 images
        if i % 5 == 0:
            df = results_to_dataframe(results)
            save_with_formatting(df, OUTPUT_FILE)
            print(f"         [Partial save: {i} images]")

    # Save final results with formatting
    df = results_to_dataframe(results)
    save_with_formatting(df, OUTPUT_FILE)

    # Print summary
    print("-" * 60)
    print("SUMMARY BY VIEW:")
    print("-" * 60)

    sorted_results = group_and_sort_results(results)
    current_view = None
    for r in sorted_results:
        if r is None:
            continue
        view_key = (r.get('location'), r.get('view_id'))
        if view_key != current_view:
            current_view = view_key
            print(f"\n{r.get('location')} - View {r.get('view_id')}:")

        marker = "  ★ BEST" if r.get('is_best') else "       "
        total = r.get('total', 'N/A')
        print(f"  {marker} {r.get('filename')}: {total}/10")

    print("\n" + "-" * 60)
    print(f"Results saved to: {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
